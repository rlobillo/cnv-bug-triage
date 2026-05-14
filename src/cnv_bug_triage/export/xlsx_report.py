"""Export triage results to a multi-sheet XLSX workbook.

Sheets:
  Run Info    — metadata (date, mode, model, filters, totals, release calendar)
  All Bugs    — all bugs with triage status
  Untriaged   — untriaged bugs with LLM suggestions
  Backport    — backport recommendations
  Duplicates  — potential duplicate pairs
  Comments    — (--apply only) comment posting status
"""

from __future__ import annotations

from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from cnv_bug_triage.runner import RunResult
    from cnv_bug_triage.schemas.config import AppConfig


_HEADER_FILL = PatternFill("solid", fgColor="1F497D")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
_TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _header_row(ws: Any, cols: list[str]) -> None:
    ws.append(cols)
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER


def _auto_width(ws: Any, min_w: int = 10, max_w: int = 60) -> None:
    for col_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[
            get_column_letter(col_cells[0].column)
        ].width = min(max_w, max(min_w, length + 2))


def _freeze_top(ws: Any) -> None:
    ws.freeze_panes = "A2"


def _shade_alt_rows(ws: Any) -> None:
    for i, row in enumerate(ws.iter_rows(min_row=2), start=1):
        if i % 2 == 0:
            for cell in row:
                cell.fill = _ALT_FILL


def _hyperlink_cell(cell: Any, url: str) -> None:
    cell.hyperlink = url
    cell.font = Font(color="0563C1", underline="single")


def _build_run_info(
    ws: Any,
    result: RunResult,
    cfg: AppConfig,
) -> None:
    ws.title = "Run Info"
    _header_row(ws, ["Field", "Value"])

    c = result.counters
    rows = [
        ("Date", result.timestamp),
        ("Mode", result.mode),
        ("Model", result.model),
        ("Run ID", result.run_id),
        ("JQL", result.jql_used),
        ("Total Bugs", str(c.total)),
        ("Triaged", str(c.triaged)),
        ("Untriaged", str(c.untriaged)),
        ("Backport Needed", str(c.backport_needed)),
        ("Duplicates Found", str(c.duplicates_found)),
    ]
    if result.mode == "apply":
        rows.append(("Comments Posted", str(c.comments_posted)))
        rows.append(("Comments Skipped", str(c.comments_skipped)))

    for rv in cfg.releases.versions:
        eol = f", EOL: {rv.eol_date}" if rv.eol_date else ""
        rows.append((
            f"Release {rv.version}",
            f"{rv.status}{eol}",
        ))

    for key, val in rows:
        ws.append([key, val])
    _auto_width(ws)
    ws.column_dimensions["B"].width = 80


def _build_all_bugs(ws: Any, result: RunResult) -> None:
    ws.title = "All Bugs"
    cols = [
        "Key", "Summary", "Status", "Type",
        "Assignee", "QA Contact", "Sprint",
        "Priority", "Fix Version", "Triaged?", "Missing Fields",
    ]
    _header_row(ws, cols)

    for r in result.results:
        bug = r.bug
        ws.append([
            bug.key,
            bug.summary,
            bug.status,
            bug.issue_type,
            bug.assignee_name or "(unassigned)",
            bug.qa_contact_name or "(unassigned)",
            bug.sprint_name or "(none)",
            bug.priority,
            ", ".join(bug.fix_versions) or "(none)",
            "Yes" if r.is_triaged else "No",
            ", ".join(r.missing_fields) or "",
        ])
        _hyperlink_cell(ws.cell(ws.max_row, 1), bug.url)

    _freeze_top(ws)
    _shade_alt_rows(ws)
    _auto_width(ws)


def _build_untriaged(ws: Any, result: RunResult) -> None:
    ws.title = "Untriaged"
    cols = [
        "Key", "Summary", "Status", "Type", "Missing Fields",
        "Suggested Priority", "Suggested Assignee",
        "Suggested QA Contact", "Suggested Sprint",
        "Suggested Fix Version", "Reasoning",
    ]
    _header_row(ws, cols)

    for r in result.results:
        if r.is_triaged:
            continue
        bug = r.bug

        def _sug(field: str) -> str:
            s = r.suggestion_for(field)
            return s.suggested_display if s else ""

        def _reason(field: str) -> str:
            s = r.suggestion_for(field)
            return s.reasoning if s else ""

        reasoning_parts = [
            f"{f}: {_reason(f)}" for f in r.missing_fields if _reason(f)
        ]

        ws.append([
            bug.key,
            bug.summary,
            bug.status,
            bug.issue_type,
            ", ".join(r.missing_fields),
            _sug("priority"),
            _sug("assignee"),
            _sug("qa_contact"),
            _sug("sprint"),
            _sug("fix_version"),
            " | ".join(reasoning_parts),
        ])
        _hyperlink_cell(ws.cell(ws.max_row, 1), bug.url)

    _freeze_top(ws)
    _shade_alt_rows(ws)
    _auto_width(ws)
    if ws.max_column >= 11:
        ws.column_dimensions[get_column_letter(11)].width = 80
        for cell in ws[get_column_letter(11)]:
            cell.alignment = _TOP_LEFT


def _build_backport(ws: Any, result: RunResult) -> None:
    ws.title = "Backport"
    cols = [
        "Bug Key", "Summary", "Fix Version",
        "Backport To", "Branch", "Urgency", "Reasoning",
    ]
    _header_row(ws, cols)

    for r in result.results:
        if not r.backport or not r.backport.recommendations:
            continue
        bug = r.bug
        fix_ver = ", ".join(bug.fix_versions) or "(none)"
        fv_sug = r.suggestion_for("fix_version")
        if fv_sug and not bug.fix_versions:
            fix_ver = f"{fv_sug.suggested_display} (suggested)"

        for rec in r.backport.recommendations:
            ws.append([
                bug.key,
                bug.summary,
                fix_ver,
                rec.version,
                rec.branch,
                rec.urgency.upper().replace("_", " "),
                rec.reasoning,
            ])
            _hyperlink_cell(ws.cell(ws.max_row, 1), bug.url)

    _freeze_top(ws)
    _shade_alt_rows(ws)
    _auto_width(ws)


def _build_duplicates(ws: Any, result: RunResult) -> None:
    ws.title = "Duplicates"
    cols = ["Bug Key", "Candidate Key", "Confidence", "Reasoning"]
    _header_row(ws, cols)

    for r in result.results:
        for dup in r.duplicates:
            ws.append([
                r.bug.key,
                dup.candidate_key,
                f"{int(dup.confidence * 100)}%",
                dup.reasoning,
            ])
            _hyperlink_cell(ws.cell(ws.max_row, 1), r.bug.url)

    _freeze_top(ws)
    _shade_alt_rows(ws)
    _auto_width(ws)


def _build_comments(ws: Any, result: RunResult) -> None:
    ws.title = "Comments"
    cols = ["Bug Key", "Status", "Reason", "Comment Text"]
    _header_row(ws, cols)

    for rec in result.comment_records:
        ws.append([rec.bug_key, rec.status, rec.reason, rec.text])

    _freeze_top(ws)
    _shade_alt_rows(ws)
    _auto_width(ws)
    if ws.max_column >= 4:
        ws.column_dimensions[get_column_letter(4)].width = 80


def build_xlsx(
    path: str,
    *,
    result: RunResult,
    cfg: AppConfig,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    _build_run_info(wb.create_sheet(), result, cfg)
    _build_all_bugs(wb.create_sheet(), result)
    _build_untriaged(wb.create_sheet(), result)
    _build_backport(wb.create_sheet(), result)
    _build_duplicates(wb.create_sheet(), result)

    if result.mode == "apply":
        _build_comments(wb.create_sheet(), result)

    wb.save(path)
