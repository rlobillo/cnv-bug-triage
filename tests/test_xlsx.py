"""Tests for export/xlsx_report.py — XLSX workbook generation."""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from cnv_bug_triage.export.xlsx_report import build_xlsx
from cnv_bug_triage.runner import CommentRecord, RunCounters, RunResult
from cnv_bug_triage.triage.checker import (
    BackportAnalysis,
    BackportRecommendation,
    DuplicateCandidate,
    TriageResult,
    TriageSuggestion,
)
from tests.conftest import failing_check, make_bug, passing_check


class TestBuildXlsx:
    def _build_and_load(self, result, cfg, tmp_path):
        path = str(tmp_path / "report.xlsx")
        build_xlsx(path, result=result, cfg=cfg)
        return load_workbook(path)

    def test_dry_run_sheet_names(self, sample_cfg, tmp_path):
        rr = RunResult(
            run_id="abc", mode="dry-run", model="gpt-4o",
            timestamp="2026-05-14", jql_used="test",
            results=[], counters=RunCounters(),
        )
        wb = self._build_and_load(rr, sample_cfg, tmp_path)
        names = wb.sheetnames
        assert "Run Info" in names
        assert "All Bugs" in names
        assert "Untriaged" in names
        assert "Backport" in names
        assert "Duplicates" in names
        assert "Comments" not in names

    def test_apply_mode_includes_comments_sheet(self, sample_cfg, tmp_path):
        rr = RunResult(
            run_id="abc", mode="apply", model="gpt-4o",
            timestamp="2026-05-14", jql_used="test",
            results=[], counters=RunCounters(),
            comment_records=[CommentRecord("CNV-1", "posted", "ok", "text")],
        )
        wb = self._build_and_load(rr, sample_cfg, tmp_path)
        assert "Comments" in wb.sheetnames

    def test_run_info_content(self, sample_cfg, tmp_path):
        counters = RunCounters(total=5, triaged=3, untriaged=2)
        rr = RunResult(
            run_id="abc123", mode="dry-run", model="gpt-4o",
            timestamp="2026-05-14 10:00 UTC", jql_used="project = X",
            results=[], counters=counters,
        )
        wb = self._build_and_load(rr, sample_cfg, tmp_path)
        ws = wb["Run Info"]
        values = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(2, ws.max_row + 1)}
        assert values["Mode"] == "dry-run"
        assert values["Model"] == "gpt-4o"
        assert values["Total Bugs"] == "5"
        assert values["Triaged"] == "3"

    def test_all_bugs_row(self, sample_cfg, tmp_path):
        bug = make_bug(key="CNV-42", summary="Test bug", priority="Major")
        checks = [passing_check(f) for f in ["assignee", "qa_contact", "sprint", "priority", "fix_version"]]
        result = TriageResult(bug=bug, checks=checks)
        rr = RunResult(
            run_id="abc", mode="dry-run", model="gpt-4o",
            timestamp="2026-05-14", jql_used="test",
            results=[result], counters=RunCounters(total=1, triaged=1),
        )
        wb = self._build_and_load(rr, sample_cfg, tmp_path)
        ws = wb["All Bugs"]
        assert ws.cell(2, 1).value == "CNV-42"
        assert ws.cell(2, 2).value == "Test bug"

    def test_untriaged_row_with_suggestions(self, sample_cfg, tmp_path):
        bug = make_bug(key="CNV-42")
        checks = [failing_check("priority")]
        sug = TriageSuggestion(
            field_name="priority", suggested_value="Major",
            suggested_display="Major", reasoning="High impact", source="llm",
        )
        result = TriageResult(bug=bug, checks=checks, suggestions=[sug])
        rr = RunResult(
            run_id="abc", mode="dry-run", model="gpt-4o",
            timestamp="2026-05-14", jql_used="test",
            results=[result], counters=RunCounters(total=1, untriaged=1),
        )
        wb = self._build_and_load(rr, sample_cfg, tmp_path)
        ws = wb["Untriaged"]
        assert ws.cell(2, 1).value == "CNV-42"
        assert ws.cell(2, 6).value == "Major"

    def test_backport_row(self, sample_cfg, tmp_path):
        bug = make_bug(key="CNV-42", fix_versions=["CNV v4.22.0"])
        checks = [passing_check("priority")]
        backport = BackportAnalysis(
            needs_backport=True,
            recommendations=[
                BackportRecommendation("4.21", "release-4.21", "must", "Critical"),
            ],
            overall_reasoning="test",
        )
        result = TriageResult(bug=bug, checks=checks, backport=backport)
        rr = RunResult(
            run_id="abc", mode="dry-run", model="gpt-4o",
            timestamp="2026-05-14", jql_used="test",
            results=[result], counters=RunCounters(total=1),
        )
        wb = self._build_and_load(rr, sample_cfg, tmp_path)
        ws = wb["Backport"]
        assert ws.cell(2, 1).value == "CNV-42"
        assert ws.cell(2, 4).value == "4.21"
        assert ws.cell(2, 6).value == "MUST"

    def test_duplicates_row(self, sample_cfg, tmp_path):
        bug = make_bug(key="CNV-42")
        checks = [failing_check("priority")]
        dup = DuplicateCandidate("CNV-99", "Similar", 0.9, "Same trace")
        result = TriageResult(bug=bug, checks=checks, duplicates=[dup])
        rr = RunResult(
            run_id="abc", mode="dry-run", model="gpt-4o",
            timestamp="2026-05-14", jql_used="test",
            results=[result], counters=RunCounters(total=1),
        )
        wb = self._build_and_load(rr, sample_cfg, tmp_path)
        ws = wb["Duplicates"]
        assert ws.cell(2, 1).value == "CNV-42"
        assert ws.cell(2, 2).value == "CNV-99"
        assert ws.cell(2, 3).value == "90%"

    def test_comments_row(self, sample_cfg, tmp_path):
        rr = RunResult(
            run_id="abc", mode="apply", model="gpt-4o",
            timestamp="2026-05-14", jql_used="test",
            results=[], counters=RunCounters(),
            comment_records=[CommentRecord("CNV-1", "posted", "ok", "Comment text")],
        )
        wb = self._build_and_load(rr, sample_cfg, tmp_path)
        ws = wb["Comments"]
        assert ws.cell(2, 1).value == "CNV-1"
        assert ws.cell(2, 2).value == "posted"
        assert ws.cell(2, 4).value == "Comment text"

    def test_release_versions_in_run_info(self, sample_cfg, tmp_path):
        rr = RunResult(
            run_id="abc", mode="dry-run", model="gpt-4o",
            timestamp="2026-05-14", jql_used="test",
            results=[], counters=RunCounters(),
        )
        wb = self._build_and_load(rr, sample_cfg, tmp_path)
        ws = wb["Run Info"]
        all_values = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        assert "Release 4.22" in all_values
        assert "Release 4.19" in all_values
